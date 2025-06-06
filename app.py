from flask import Flask, render_template, Response, jsonify, redirect, url_for, request, flash
from datetime import datetime, timedelta, time, date
from dotenv import dotenv_values
from openpyxl import load_workbook
import io, os, csv, socket, time, mysql.connector, requests, logging, re
from zk import ZK, const

config = dotenv_values(".env")

db_config = {
    "host": config["DB_HOST"],
    "user": config["DB_USER"],
    "password": config["DB_PASSWORD"],
    "database": config["DB_NAME"]
}

PORT = 4370
TIMEOUT = .5

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY')

terminals = [
    {"name": "Terminal 1", "ip": "192.168.2.221", 'port': 4370},
    {"name": "Terminal 2", "ip": "192.168.2.222", 'port': 4370},
]

cyrillic_to_latin = str.maketrans({
'А': 'A', 'а': 'a',
'Б': 'B', 'б': 'b',
'В': 'V', 'в': 'v',
'Г': 'G', 'г': 'g',
'Д': 'D', 'д': 'd',
'Е': 'E', 'е': 'e',
'Ж': 'Zh', 'ж': 'zh',
'З': 'Z', 'з': 'z',
'И': 'I', 'и': 'i',
'Й': 'Y', 'й': 'y',
'К': 'K', 'к': 'k',
'Л': 'L', 'л': 'l',
'М': 'M', 'м': 'm',
'Н': 'N', 'н': 'n',
'О': 'O', 'о': 'o',
'П': 'P', 'п': 'p',
'Р': 'R', 'р': 'r',
'С': 'S', 'с': 's',
'Т': 'T', 'т': 't',
'У': 'U', 'у': 'u',
'Ф': 'F', 'ф': 'f',
'Х': 'H', 'х': 'h',
'Ц': 'Ts', 'ц': 'ts',
'Ч': 'Ch', 'ч': 'ch',
'Ш': 'Sh', 'ш': 'sh',
'Щ': 'Sht', 'щ': 'sht',
'Ъ': 'A', 'ъ': 'a',
'Ь': '', 'ь': '',
'Ю': 'Yu', 'ю': 'yu',
'Я': 'Ya', 'я': 'ya',
})

logging.basicConfig(level=logging.INFO)

UPLOAD_FOLDER = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'uploads')
ALLOWED_EXTENSIONS = {'csv', 'xlsx'}

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['ALLOWED_EXTENSIONS'] = ALLOWED_EXTENSIONS

# Function to check allowed file extensions (CSV or XLSX)
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def get_card_numbers_from_zkteco():
    api_url = terminals
    response = requests.get(api_url)
    
    if response.status_code == 200:
        return response.json()
    else:
        return []

def fetch_logs_from_db(start_date=None, end_date=None):
    try:
        conn = mysql.connector.connect(**db_config)
        
        # Using 'with' to ensure cursor and connection are properly closed
        with conn.cursor(dictionary=True) as cursor:
            query = "SELECT userID, employee, workday, clockIn, clockOut FROM logs"
            filters = []
            params = []

            if start_date:
                if isinstance(start_date, str):
                    start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
                filters.append("workday >= %s")
                params.append(start_date)

            if end_date:
                if isinstance(end_date, str):
                    end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
                filters.append("workday <= %s")
                params.append(end_date)

            if filters:
                query += " WHERE " + " AND ".join(filters)

            # Execute the query with parameters to prevent SQL injection
            cursor.execute(query, tuple(params))
            logs = cursor.fetchall()

        conn.close()

        return logs
    
    except mysql.connector.Error as e:
        print(f"Database error: {e}")
        return {"error": "Database connection failed."}
    except Exception as e:
        print(f"Error: {e}")
        return {"error": "An unexpected error occurred."}

def check_terminal_status(ip):
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(2)
        result = sock.connect_ex((ip, PORT))
        sock.close()
        
        if result == 0:
            zk = ZK(ip, port=PORT, timeout=TIMEOUT)
            conn = zk.connect()
            if conn:
                conn.disconnect()
                return True
        return False
    except Exception as e:
        print(f"Error checking terminal status for {ip}: {e}")
        return False  # If an error occurs, mark it as offline

def get_connection(ip):
    try:
        zk = ZK(ip, port=PORT, timeout=TIMEOUT, password=0, force_udp=False, ommit_ping=False)
        conn = zk.connect()
        return conn
    except Exception as e:
        print(f"Error connecting to {ip}: {e}")
        return None
    
def fetch_attendance_from_terminal(ip, label):
    zk = ZK(ip, port=PORT, timeout=TIMEOUT, password=0, force_udp=False, ommit_ping=True)
    logs = []
    try:
        conn = zk.connect()
        conn.disable_device()
        attendance = conn.get_attendance()
        users = conn.get_users()
        user_dict = {user.user_id: user.name for user in users}

        for log in attendance:
            status = "Unknown Status"
            log_status = str(log.status).strip()

            if log_status == "1":
                status = "Clocked In"
            elif log_status == "2":
                status = "Clocked Out"
            
            logs.append({
                "terminal": label,
                "user_id": log.user_id,
                "username": user_dict.get(log.user_id, "Unknown"),
                "timestamp": log.timestamp,
                "status": status,
            })

        conn.enable_device()
        conn.disconnect()
    except Exception as e:
        print(f"Error connecting to {label} ({ip}):", e)
        logs.append({
            "terminal": label,
            "user_id": "-",
            "username": "Device Offline",
            "timestamp": "-",
            "status": "Offline"
        })
    return logs

def clear_all_users_from_device(ip):
    zk = ZK(ip, port=PORT, timeout=5, force_udp=True)
    try:
        conn = zk.connect()
        conn.disable_device()
        users = conn.get_users()
        
        for user in users:
            if user.uid > 1:  # Avoid deleting admin user (UID 0 or 1)
                conn.delete_user(uid=user.uid)

        conn.enable_device()
        conn.disconnect()
        print(f"[INFO] Cleared users from {ip}")
        return True
    except Exception as e:
        print(f"[ERROR] Could not clear users from {ip}: {e}")
        return False
    
def clear_logs_from_device(ip):
    zk = ZK(ip, port=PORT, timeout=5, force_udp=True)
    try:
        conn = zk.connect()
        conn.disable_device()
        conn.clear_attendance()  # Clears all attendance logs
        conn.enable_device()
        conn.disconnect()
        print(f"[INFO] Logs cleared from {ip}")
        return True
    except Exception as e:
        print(f"[ERROR] Could not clear logs from {ip}: {e}")
        return False


@app.route('/api/ping', methods=['POST'])
def api_ping():
    ip = request.json.get('ip')
    conn = get_connection(ip)
    if conn:
        try:
            conn.disconnect()
            return jsonify({"success": True, "message": "Ping successful"})
        except:
            return jsonify({"success": False, "message": "Ping failed"})
    return jsonify({"success": False, "message": "Device offline"})


@app.route('/api/reboot', methods=['POST'])
def reboot_terminal():
    try:
        data = request.get_json()  # Get the JSON data from the request body
        ip = data.get("ip")  # Extract the IP address from the data

        if not ip:
            return jsonify({"success": False, "message": "IP address is required."}), 400

        # Check if the terminal is online (ping it)
        if not check_terminal_status(ip):
            return jsonify({"success": False, "message": "Device is offline, unable to reboot."}), 400

        # Proceed with the reboot process if the device is online
        conn = get_connection(ip)
        if conn:
            try:
                conn.restart()  # Restart the terminal
                return jsonify({"success": True, "message": f"Device at {ip} restarted successfully."})
            except Exception as e:
                print(f"Error restarting device {ip}: {e}")
                return jsonify({"success": False, "message": f"Failed to restart device at {ip}."}), 500
        else:
            return jsonify({"success": False, "message": f"Unable to connect to {ip}."}), 500

    except Exception as e:
        print(f"Error rebooting device {ip}: {e}")
        return jsonify({"success": False, "message": "Failed to restart device."}), 500


@app.route('/api/set_time', methods=['POST'])
def api_set_time():
    ip = request.json.get('ip')
    conn = get_connection(ip)
    if conn:
        try:
            conn.set_time(datetime.now())
            conn.disconnect()
            return jsonify({"success": True, "message": "Time synced successfully"})
        except:
            return jsonify({"success": False, "message": "Set time failed"})
    return jsonify({"success": False, "message": "Device offline"})


@app.route('/api/get_time', methods=['POST'])
def api_get_time():
    ip = request.json.get('ip')
    conn = get_connection(ip)
    if conn:
        try:
            current_time = conn.get_time()
            conn.disconnect()
            return jsonify({"success": True, "time": current_time.strftime('%Y-%m-%d %H:%M:%S')})
        except:
            return jsonify({"success": False, "message": "Failed to get time"})
    return jsonify({"success": False, "message": "Device offline"})


@app.route('/api/refresh', methods=['POST'])
def api_refresh():
    return jsonify({"success": True})  # We’ll reload the page on frontend instead


@app.route('/')
def index():
    all_logs = []
    enriched_terminals = []
    current_terminal_time = "Unavailable"

    for terminal in terminals:
        ip = terminal["ip"]

        status = "On" if check_terminal_status(ip) else "Off"
        user_count = 0
        current_time = "Unavailable"

        try:
            if status == "On":
                conn = get_connection(ip)
                if conn:
                    user_count = len(conn.get_users())
                    current_time = conn.get_time().strftime('%Y-%m-%d %H:%M:%S')  # Get the time from the terminal
                    current_terminal_time = current_time  # Save the first available terminal time
                    conn.disconnect()
        except Exception as e:
            print(f"Error with {terminal['name']} at {ip}:", e)
            status = "Off"

        enriched_terminals.append({
            "name": terminal["name"],
            "ip": ip,
            "status": status,
            "user_count": user_count,
            "current_time": current_time
        })

        if status == "On":
            logs = fetch_attendance_from_terminal(ip, terminal["name"])
            all_logs.extend(logs)
        else:
            all_logs.append({
                "terminal": terminal["name"],
                "user_id": "-",
                "username": "Терминала е офлайн",
                "timestamp": "-",
                "status": "Офлайн"
            })

    all_logs.sort(key=lambda x: x['timestamp'] if x['timestamp'] != "-" else datetime.min, reverse=True)

    return render_template('index.html', logs=all_logs, terminals=enriched_terminals, terminal_time=current_terminal_time)


@app.route('/export_csv')
def export_csv():
    all_logs = []
    for terminal in terminals:
        logs = fetch_attendance_from_terminal(terminal["ip"], terminal["name"])
        all_logs.extend(logs)

    fieldnames = ['terminal', 'user_id', 'username', 'timestamp', 'status']

    def generate():
        output = io.StringIO(newline='')
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()

        for log in all_logs:
            log['timestamp'] = log['timestamp'].strftime('%Y-%m-%d %H:%M:%S') if hasattr(log['timestamp'], 'strftime') else log['timestamp']
            writer.writerow(log)

        csv_string = output.getvalue()
        output.close()

        # Add UTF-8 BOM so Excel opens it correctly
        bom = '\ufeff'  # UTF-8 Byte Order Mark
        return bom + csv_string

    return Response(
        generate(),
        mimetype='text/csv',
        headers={"Content-Disposition": "attachment; filename=attendance_logs.csv"}
    )

@app.route('/export_json')
def export_json():
    all_logs = []
    for terminal in terminals:
        logs = fetch_attendance_from_terminal(terminal["ip"], terminal["name"])
        all_logs.extend(logs)

    # Convert datetime objects to strings
    for log in all_logs:
        if hasattr(log['timestamp'], 'strftime'):
            log['timestamp'] = log['timestamp'].strftime('%Y-%m-%d %H:%M:%S')

    return jsonify(all_logs)

@app.route('/attendance', methods=['GET', 'POST'])
def attendance():
    start_date = request.args.get('start_date', None)
    end_date = request.args.get('end_date', None)

    # Fetch logs from the database
    all_logs = fetch_logs_from_db(start_date, end_date)

    # Convert string dates to datetime objects if provided
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()  # Convert to datetime.date
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()  # Convert to datetime.date

    # Filter logs by date range if start_date and end_date are provided
    if start_date or end_date:
        filtered_logs = []
        for log in all_logs:
            # Check if log['workday'] is already a datetime.date object
            log_time = log['workday'] if isinstance(log['workday'], date) else datetime.strptime(log['workday'], '%Y-%m-%d').date()

            if start_date and log_time < start_date:
                continue
            if end_date and log_time > end_date:
                continue

            filtered_logs.append(log)
        all_logs = filtered_logs

    # Group by user_id and get first clock-in and last clock-out
    user_logs = {}
    for log in all_logs:
        user_id = log['userID']
        workday = log['workday'] if isinstance(log['workday'], date) else datetime.strptime(log['workday'], '%Y-%m-%d').date()
        key = (user_id, workday)  # Grouping key

        timestamp_in = None
        timestamp_out = None

        # Convert clock-in and clock-out to datetime objects if they are time or timedelta
        if isinstance(log['clockIn'], datetime):
            timestamp_in = datetime.combine(datetime.today(), log['clockIn']) if log['clockIn'] else None
        elif isinstance(log['clockIn'], timedelta):
            timestamp_in = datetime.combine(datetime.today(), datetime.min.time()) + log['clockIn'] if log['clockIn'] else None

        if isinstance(log['clockOut'], datetime):
            timestamp_out = datetime.combine(datetime.today(), log['clockOut']) if log['clockOut'] else None
        elif isinstance(log['clockOut'], timedelta):
            timestamp_out = datetime.combine(datetime.today(), datetime.min.time()) + log['clockOut'] if log['clockOut'] else None

        if key not in user_logs:
            user_logs[key] = {
                'employee': log['employee'],
                'first_clock_in': None,
                'last_clock_out': None,
                'workday': workday
            }

        # Record earliest clock-in
        if timestamp_in:
            if user_logs[key]['first_clock_in'] is None or timestamp_in < user_logs[key]['first_clock_in']:
                user_logs[key]['first_clock_in'] = timestamp_in

        # Record latest clock-out
        if timestamp_out:
            if user_logs[key]['last_clock_out'] is None or timestamp_out > user_logs[key]['last_clock_out']:
                user_logs[key]['last_clock_out'] = timestamp_out

    # Prepare the list to be displayed in the table
    display_logs = []
    for (user_id, workday), user_data in user_logs.items():
        display_logs.append({
            "id": user_id,
            "user": user_data['employee'],
            "date": user_data['workday'],
            "first_clock_in": user_data['first_clock_in'].strftime('%H:%M') if user_data['first_clock_in'] else "Not Clocked In",
            "last_clock_out": user_data['last_clock_out'].strftime('%H:%M') if user_data['last_clock_out'] else "Not Clocked Out"
        })

    # Sort logs by date descending
    display_logs.sort(key=lambda x: x['date'], reverse=True)

    return render_template('attendance.html', logs=display_logs, start_date=start_date, end_date=end_date)

@app.route('/scan_card', methods=['POST'])
def scan_card():
    try:
        # Get the data sent by the terminal (from the scan event)
        data = request.get_json()

        # Extract necessary information
        user_id = data['userID']
        employee_name = data['employee']
        clock_in = data.get('clockIn')  # Assuming clock-in is provided
        clock_out = data.get('clockOut')  # If applicable

        # Ensure required fields are present
        if not user_id or not employee_name or not clock_in:
            return jsonify({"error": "Missing required fields."}), 400

        # Get the current date (workday)
        workday = datetime.today().date()

        # Connect to the database
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()

        # Insert log into the database
        insert_query = """
        INSERT INTO logs (userID, employee, workday, clockIn, clockOut)
        VALUES (%s, %s, %s, %s, %s)
        """
        cursor.execute(insert_query, (user_id, employee_name, workday, clock_in, clock_out))

        # Commit and close connection
        conn.commit()
        cursor.close()
        conn.close()

        return jsonify({"success": "Log added successfully."}), 200

    except mysql.connector.Error as e:
        print(f"Database error: {e}")
        return jsonify({"error": "Database connection failed."}), 500
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"error": "An unexpected error occurred."}), 500

@app.route('/export_attendance', methods=['GET'])
def export_attendance():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    all_logs = fetch_logs_from_db(start_date, end_date)

    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

    if start_date or end_date:
        filtered_logs = []
        for log in all_logs:
            log_time = log['workday'] if isinstance(log['workday'], date) else datetime.strptime(log['workday'], '%Y-%m-%d').date()
            if start_date and log_time < start_date:
                continue
            if end_date and log_time > end_date:
                continue
            filtered_logs.append(log)
        all_logs = filtered_logs

    user_logs = {}
    for log in all_logs:
        user_id = log['userID']
        workday = log['workday'] if isinstance(log['workday'], date) else datetime.strptime(log['workday'], '%Y-%m-%d').date()
        key = (user_id, workday)

        timestamp_in = None
        timestamp_out = None

        if isinstance(log['clockIn'], time):
            timestamp_in = datetime.combine(datetime.today(), log['clockIn'])
        elif isinstance(log['clockIn'], timedelta):
            timestamp_in = datetime.combine(datetime.today(), datetime.min.time()) + log['clockIn']

        if isinstance(log['clockOut'], time):
            timestamp_out = datetime.combine(datetime.today(), log['clockOut'])
        elif isinstance(log['clockOut'], timedelta):
            timestamp_out = datetime.combine(datetime.today(), datetime.min.time()) + log['clockOut']

        if key not in user_logs:
            user_logs[key] = {
                'employee': log['employee'],
                'first_clock_in': None,
                'last_clock_out': None,
                'workday': workday
            }

        if timestamp_in:
            if user_logs[key]['first_clock_in'] is None or timestamp_in < user_logs[key]['first_clock_in']:
                user_logs[key]['first_clock_in'] = timestamp_in

        if timestamp_out:
            if user_logs[key]['last_clock_out'] is None or timestamp_out > user_logs[key]['last_clock_out']:
                user_logs[key]['last_clock_out'] = timestamp_out

    export_data = []
    for (user_id, workday), user_data in user_logs.items():
        export_data.append({
            "User ID": user_id,
            "Employee": user_data['employee'],
            "Date": user_data['workday'].strftime('%Y-%m-%d'),
            "First Clock-In": user_data['first_clock_in'].strftime('%H:%M:%S') if user_data['first_clock_in'] else "Not Clocked In",
            "Last Clock-Out": user_data['last_clock_out'].strftime('%H:%M:%S') if user_data['last_clock_out'] else "Not Clocked Out"
        })

    def generate_csv():
        output = io.StringIO(newline='')
        writer = csv.DictWriter(output,
                                fieldnames=["User ID", "Employee", "Date", "First Clock-In", "Last Clock-Out"],
                                delimiter=';')  # <<< Това е ключовото
        writer.writeheader()
        for row in export_data:
            writer.writerow(row)
        return output.getvalue()

    return Response(
        generate_csv(),
        mimetype='text/csv',
        headers={"Content-Disposition": "attachment; filename=attendance_log.csv"}
    )


@app.route('/users')
def users_page():
    all_users = []
    
    try:
        # Connect to the MySQL database using the db_config dictionary
        conn = mysql.connector.connect(**db_config)
        if conn is None:
            print("Failed to connect to the database.")  # Debugging log
            return render_template('users.html', users=all_users)

        cursor = conn.cursor(dictionary=True)  # Using dictionary cursor to fetch rows as dictionaries

        # Query to fetch user data from the database
        query = "SELECT ID, employeeName, cardNumber FROM employees ORDER BY ID ASC"
        cursor.execute(query)

        # Fetch all users from the database
        users_from_db = cursor.fetchall()

        if users_from_db:
            print(f"Fetched {len(users_from_db)} users from the database.")  # Debugging log
        else:
            print("No users found in the database.")  # Debugging log

        for user in users_from_db:
            card = str(user['cardNumber'])  # Ensure card is treated as a string
            card = "N/A" if card == "0" else card  # Replace '0' with 'N/A'

            # Add user data to the list
            all_users.append({
                'user_id': user['ID'],
                'name': user['employeeName'].translate(cyrillic_to_latin),
                'card': card
            })

        # Close the database connection
        cursor.close()
        conn.close()

    except Exception as e:
        print(f"Failed to fetch users: {e}")

    return render_template('users.html', users=all_users)

# Route to add employees or upload CSV/XLSX
@app.route('/add-employees', methods=['GET', 'POST'])
def add_employee():
    if request.method == 'POST':
        # Check if a file was uploaded
        if 'file' not in request.files:
            return "No file part"

        file = request.files['file']

        # If no file is selected or file is empty
        if file.filename == '':
            return "No selected file"

        # If the file is allowed
        if file and allowed_file(file.filename):
            try:
                # Save the uploaded file
                filename = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
                file.save(filename)

                # Check the file extension and handle accordingly
                if filename.endswith('.csv'):
                    users_to_add = []
                    # Read and process CSV
                    with open(filename, mode='r') as f:
                        csv_reader = csv.DictReader(f)
                        for row in csv_reader:
                            user_id = row['userID']
                            employee_name = row['name'].translate(cyrillic_to_latin)
                            card_number = row['cardNumber']
                            users_to_add.append((user_id, employee_name.translate(cyrillic_to_latin), card_number))

                    # Insert users into the database (for CSV)
                    conn = mysql.connector.connect(**db_config)
                    cursor = conn.cursor()
                    query = "INSERT INTO employees (ID, employeeName, cardNumber) VALUES (%s, %s, %s)"
                    cursor.executemany(query, users_to_add)
                    conn.commit()
                    cursor.close()
                    conn.close()

                elif filename.endswith('.xlsx'):
                    users_to_add = []
                    # Load the Excel workbook
                    wb = load_workbook(filename)
                    sheet = wb.active

                    # Iterate over the rows in the Excel sheet (skip the header row)
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        user_id, employee_name, card_number = row
                        users_to_add.append((user_id, employee_name.translate(cyrillic_to_latin), card_number))

                    # Insert users into the database (for XLSX)
                    conn = mysql.connector.connect(**db_config)
                    cursor = conn.cursor()
                    query = "INSERT INTO employees (ID, employeeName, cardNumber) VALUES (%s, %s, %s)"
                    cursor.executemany(query, users_to_add)
                    conn.commit()
                    cursor.close()
                    conn.close()

                return redirect(url_for('users_page'))  # Redirect to the user list page after successful upload

            except Exception as e:
                print(f"Error: {e}")
                return "There was an error processing the file."
        
        else:
            return "Invalid file format. Please upload a CSV or XLSX file."

    return render_template('add_employee.html')  # Render the form to upload file if method is GET

@app.route('/upload-employees', methods=['GET'])
def upload_employees_to_terminals():
    status_messages = []
    overall_success = True  # Flag to track overall success
    # Function to clear all users from the terminal
    def clear_all_users_from_device(ip):
        try:
            zk = ZK(ip, port=PORT, timeout=5)
            conn = zk.connect()
            if conn is None:
                raise Exception(f"Could not connect to terminal {ip}")

            conn.disable_device()

            # Fetch existing users from the terminal
            users = conn.get_users()
            if not users:
                print(f"No users found on terminal {ip}")
            else:
                print(f"Deleting users from terminal {ip}...")
                for user in users:
                    uid = user.uid  # Access the UID of the user object
                    print(f"Deleting user UID={uid} from terminal {ip}")
                    success = conn.delete_user(uid)
                    if success:
                        print(f"User UID={uid} deleted successfully.")
                    else:
                        print(f"Failed to delete user UID={uid}.")

            conn.enable_device()
            conn.disconnect()
            return True

        except Exception as e:
            print(f"Error deleting users from terminal {ip}: {str(e)}")
            return False

    try:
        # Step 1: Fetch employees from the database
        db = mysql.connector.connect(**db_config)
        cursor = db.cursor(dictionary=True)
        cursor.execute("SELECT ID, employeeName, cardNumber FROM employees")
        employees = cursor.fetchall()
        cursor.close()
        db.close()

        # Status messages to store results
        status_messages = []

        # Step 2: Iterate through each terminal
        for device in terminals:
            zk = ZK(device['ip'], port=device['port'], timeout=5)
            try:
                # Clear all users from the terminal before uploading new users
                print(f"Attempting to clear existing users from {device['ip']}...")
                success = clear_all_users_from_device(device['ip'])
                if not success:
                    status_messages.append(f"Failed to clear users from {device['name']} ({device['ip']})")
                    continue  # Skip to next terminal if user deletion fails

                # Step 3: Connect to the terminal and upload new users
                conn = zk.connect()
                conn.disable_device()

                for emp in employees:
                    try:
                        # Sanitize and validate input
                        user_id = str(emp.get('ID'))
                        uid = int(emp.get('ID'))

                        name = emp.get('employeeName').translate(cyrillic_to_latin) or 'Unknown'
                        name = re.sub(r'[^\x00-\x7F]', '', name)[:24]

                        card = str(emp.get('cardNumber') or '')
                        if card == "0" or not card.isdigit():
                            card = ''

                        # Print debug info (optional)
                        print(f"Uploading: UID={uid}, user_id={user_id}, name={name}, card={card}")

                        conn.set_user(
                            uid=uid,
                            name=name.translate(cyrillic_to_latin),
                            privilege=const.USER_DEFAULT,
                            password='',
                            group_id='',
                            user_id=user_id,
                            card=card
                        )

                    except Exception as user_err:
                        msg = f"Error uploading user {emp.get('employeeName')}: {str(user_err)}"
                        print(msg)
                        status_messages.append(msg)

                conn.enable_device()
                conn.disconnect()
                status_messages.append(f"✔ Успешно добавени потребители към {device['name']}.")

            except Exception as e:
                status_messages.append(f"Error connecting to terminal {device['ip']}: {str(e)}")

        # Flash all status messages
        for msg in status_messages:
            flash(msg, 'info')

    except Exception as e:
        flash(f'Error accessing the database: {str(e)}', 'danger')

    return redirect(url_for('users_page'))

@app.route('/fetch-logs', methods=['POST'])
def fetch_logs():
    try:
        # Connect to the database
        db = mysql.connector.connect(**db_config)
        cursor = db.cursor()

        # Fetch logs from each terminal
        for device in terminals:
            zk = ZK(device['ip'], port=device['port'], timeout=5)
            try:
                # Connect to the terminal
                conn = zk.connect()
                if conn is None:
                    raise Exception(f"Could not connect to terminal {device['ip']}")

                # Fetch logs from the terminal
                logs = conn.get_attendance()

                if logs:
                    # Process and insert logs into the database
                    for log in logs:
                        user_id = log.user_id

                        # Handle timestamp conversion
                        timestamp = log.timestamp  # Assuming log.timestamp is a datetime object
                        if isinstance(timestamp, str):  # If it's a string, convert it
                            timestamp = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')
                        
                        workday = timestamp.date()  # Extract workday (date) from the timestamp
                        clock_in = timestamp.time()  # Extract clock-in time from the timestamp
                        clock_out = None  # You can update this later if needed

                        # Fetch employee name based on user_id
                        cursor.execute("SELECT employeeName FROM employees WHERE ID = %s", (user_id,))
                        employee_name = cursor.fetchone()
                        if employee_name:
                            employee = employee_name[0]
                        else:
                            employee = 'Unknown'

                        # Insert log into the database
                        cursor.execute(
                            """
                            INSERT INTO logs (userID, employee, workday, clockIn, clockOut)
                            VALUES (%s, %s, %s, %s, %s)
                            """,
                            (user_id, employee, workday, clock_in, clock_out)
                        )
                    db.commit()
                    print(f"Logs successfully fetched from terminal {device['ip']}.")
                else:
                    print(f"No logs found on terminal {device['ip']}.")
                
                conn.disconnect()

            except Exception as e:
                print(f"Error fetching logs from terminal {device['ip']}: {str(e)}")

        # Close the database connection
        cursor.close()
        db.close()

        # Flash a success message and redirect
        flash('Logs successfully fetched and saved to the database.', 'success')

    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')

    return redirect(url_for('attendance'))


@app.route('/delete-all-employees', methods=['POST'])
def delete_all_employees():
    try:
        # Step 1: Clear from the database
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()

        # Delete all employees from the database
        cursor.execute("DELETE FROM employees")
        conn.commit()

        cursor.close()
        conn.close()

        return redirect(url_for('users_page'))

    except Exception as e:
        return f"Error: {str(e)}", 500


@app.route('/delete-logs', methods=['POST'])
def delete_logs():
    try:
        # Step 1: Clear logs from each terminal
        for terminal in terminals:
            success = clear_logs_from_device(terminal["ip"])
            if not success:
                return f"Failed to clear logs from {terminal['name']} ({terminal['ip']})", 500

        return redirect(url_for('index'))  # Or any page you prefer after logs are deleted

    except Exception as e:
        return f"Error: {str(e)}", 500

@app.route('/delete-db-log', methods=['GET', 'POST'])
def delete_db_log():
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()
        cursor.execute("DELETE FROM logs")
        conn.commit()
        cursor.close()
        conn.close()

        return redirect(url_for('attendance'))
    
    except Exception as e:
        return f"Error: {str(e)}", 500

if __name__ == '__main__':
    app.run(debug=True)